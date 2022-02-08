from ScopeFoundry import Measurement
import datetime
import numpy as np
import random
import pickle
import time
import os
import logging
from Daq_Functions.DAQSimpleDOTask import DAQSimpleDOTask
from Daq_Functions.DAQSimpleDITask import DAQSimpleDITask
from Classes.Define_Odors import OdorGen
from openpyxl import Workbook

# water freely delivered
print('opening water valve')
self.waterR.high()
d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
d = self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[0])
self.check_licking_1spout(self.duration_water_large)
print('closing water valve')
self.waterR.low()
d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
d = self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[1])

# water triggered by licking
