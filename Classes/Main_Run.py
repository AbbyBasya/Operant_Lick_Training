from ScopeFoundry import Measurement
from ScopeFoundry.helper_funcs import sibling_path, load_qt_ui_file
import datetime
import numpy as np
import random
import pickle
import time
import os
import logging
from Daq_Functions.DAQSimpleDOTask import DAQSimpleDOTask
from Daq_Functions.DAQSimpleDITask import DAQSimpleDITask
from Classes.Define_Odors import OdorGen
from openpyxl import Workbook
from Classes.Performance_Stats import StatRec





class AbbyTraining(Measurement):

    def __init__(self):

        self.mouse = '3'

        self.phase = 'alternating blocks go_nogo'
        self.condition = 'Operant'
        self.punishment = 'Punish'
        self.numtrials = 20
        self.number_of_blocks = 10

        # list of odors
        self.list = [6, 7]

        # self.events_path = "C:/Abby Behavior Data/self.mouse/Feb2225/". format(self.condition, self.mouse)
        self.events_path = "C:/Abby Behavior Data/Pilot2/experiment_data_2022_1_{0}/{1}/".format(self.phase, self.mouse)
        self.events_filename = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M") + '{}.xlsx'.format(self.phase)
        self.odor_index = [1, 0]  # odor list index POSITION 0 is reward,  1 is unrewarded, 2 is control odors

        self.allowedExtraLicks = 0

        if self.punishment == 'Punish':
            self.punish = True

        if self.condition == 'Operant':
            self.operant = True
            self.licknum = 1
            self.baseline = 0
            # make a baseline licks?
        else:
            self.operant = False
            self.licknum = 0

        # pre-training
        self.p_go = 0.4
        self.p_reward_go = 0.75
        self.p_no_go = 0.2
        self.p_empty = 0.4
        self.p_reward_empty = 0.75

        # if self.phase == 'free_water':
        # self.p_reward_empty = 0
        # else:
        #  self.p_reward_empty = 0.75

        self.counter = np.zeros(9)

        self.duration_rec_on_before = 2
        self.duration_odor_on = 2
        self.duration_odor_to_action = 0
        self.duration_action_window = 5  # 2.5
        self.duration_water_large = 0.04  # 0.05
        self.duration_rec_on_after = 4
        self.duration_ITI = np.random.exponential(2, size=(
                self.number_of_blocks * (self.numtrials + int(round((self.numtrials) / 2)))))

        self.waterline = 3
        self.filename = self.events_path + self.events_filename

    def setup(self):

        self.ui_filename = sibling_path('C:/Abby Behavior Data/self.mouse/Feb2225', "lick_training_plot.ui")

        self.ui=load_qt_ui_file(self.ui_filename)

        self.settings.New('save_h5', dtype=bool, initial=False)
        self.settings.New('train', dtype=bool, initial=False, ro=False)
        self.settings.New('record_calcium', dtype=bool, initial=False, ro=False)
        self.settings.New('block_number', dtype=int, initial=10)
        self.settings.New('save_movie', dtype=bool, initial=False, ro=False)
        self.settings.New('movie_on', dtype=bool, initial=False, ro=True)
        self.settings.New('random', dtype=bool, initial=False, ro=False)
        self.settings.New('high_chance', dtype=float, initial=0.5, vmin=0, vmax=1)
        self.settings.New('audio_on', dtype=bool, initial=True, ro=False)
        self.settings.New('lick_training', dtype=bool, initial=False)
        self.settings.New('free_drop', dtype=bool, initial=False)
        self.settings.New('sniff_lock', dtype=bool, initial=False)
        self.settings.New('threshold', dtype=float, initial=0.6)
        self.settings.New('clean_level', dtype=int, initial=94)
        self.settings.New('frame_count', dtype=int, initial=0, ro=True)
        self.settings.New('inference_dropped_frame', dtype=int, initial=0, ro=True)
        self.settings.New('cv_lick', dtype=bool, initial=False)
        self.settings.New('inferred_lick', dtype=int, initial=0, vmin=0, vmax=2, ro=True)
        self.settings.New('model_path', dtype='file', is_dir=False,
                          initial='D:\\Hao\\VOTA\\VOTA_Control\\VOTAScopeMS\\trained_model_alex_small.pickle')
        '''
        setting up experimental setting parameters for task
        '''
        exp_settings = []

        exp_settings.append(self.settings.New('block', dtype=int, initial=5))
        exp_settings.append(self.settings.New('forced_side', dtype=int, initial=0, vmin=0, vmax=2))
        exp_settings.append(self.settings.New('delay', dtype=int, initial=1200, vmin=0))
        exp_settings.append(self.settings.New('go', dtype=int, initial=1500))
        exp_settings.append(self.settings.New('refract', dtype=int, initial=3500, vmin=0))
        exp_settings.append(self.settings.New('punish', dtype=int, initial=10000))

        exp_settings.append(self.settings.New('channel1', dtype=int, initial=6))
        exp_settings.append(self.settings.New('channel2', dtype=int, initial=6))
        exp_settings.append(self.settings.New('level1', dtype=int, initial=100, vmin=0, vmax=100))
        exp_settings.append(self.settings.New('level2', dtype=int, initial=100, vmin=0, vmax=100))
        exp_settings.append(self.settings.New('Tpulse1', dtype=int, initial=50))
        exp_settings.append(self.settings.New('Tpulse2', dtype=int, initial=20))
        exp_settings.append(self.settings.New('interval1', dtype=int, initial=120))
        exp_settings.append(self.settings.New('interval2', dtype=int, initial=1000))
        exp_settings.append(self.settings.New('kernel1', dtype=int, initial=0, vmin=0, vmax=10000))
        exp_settings.append(self.settings.New('kernel2', dtype=int, initial=250, vmin=0, vmax=10000))

        self.exp_settings = exp_settings
        '''
        Setting up lqs for recording stats
        '''
        self.stat_settings = []
        self.stat_settings.append(self.settings.New('trial', dtype=int, initial=0, ro=True))
        self.stat_settings.append(self.settings.New('success', dtype=int, initial=0, ro=True))
        self.stat_settings.append(self.settings.New('failure', dtype=int, initial=0, ro=True))
        self.stat_settings.append(self.settings.New('early', dtype=int, initial=0, ro=True))
        self.stat_settings.append(self.settings.New('idle', dtype=int, initial=0, ro=True))
        self.stat_settings.append(self.settings.New('success_percent', dtype=int, initial=0, ro=True))
        self.stat_settings.append(self.settings.New('failure_percent', dtype=int, initial=0, ro=True))
        self.stat_settings.append(self.settings.New('early_percent', dtype=int, initial=0, ro=True))
        self.stat_settings.append(self.settings.New('idle_percent', dtype=int, initial=0, ro=True))

        '''
        Setting up lqs for indicator lights
        '''
        self.state_ind = []
        self.reward_ind = []
        self.lick_ind = []

        self.state_ind.append(self.settings.New('delay_ind', dtype=bool, initial=False, ro=True))
        self.state_ind.append(self.settings.New('go_ind', dtype=bool, initial=False, ro=True))
        self.state_ind.append(self.settings.New('refract_ind', dtype=bool, initial=False, ro=True))
        self.state_ind.append(self.settings.New('punish_ind', dtype=bool, initial=False, ro=True))


        self.all_ind = self.state_ind + self.reward_ind + self.lick_ind

        # self.settings.New('sampling_period', dtype=float, unit='s', initial=0.005)

        # Create empty numpy array to serve as a buffer for the acquired data
        # self.buffer = np.zeros(10000, dtype=float)

        # Define how often to update display during a run
        self.display_update_period = 0.04
        '''
        add reference to hardware
        '''
        # Convenient reference to the hardware used in the measurement
        self.daq_ai = self.app.hardware['daq_ai']
        # self.arduino_sol = self.app.hardware['arduino_sol']
        self.water = self.app.hardware['arduino_water']
        self.camera = self.app.hardware['top_cam']
        self.sound = self.app.hardware['sound']
        # self.motor = self.app.hardware['arduino_motor']
        self.kernel = np.zeros(250)
        # self.micro_cam = self.app.hardware['micro_cam']
        self.recorder = self.app.hardware['flirrec']
        # self.daq_do = self.app.hardware['daq_do']

    def setup_figure(self):
        """
                Runs once during App initialization, after setup()
                This is the place to make all graphical interface initializations,
                build plots, etc.
                """

        # connect ui widgets to measurement/hardware settings or functions
        self.ui.start_pushButton.clicked.connect(self.start)
        self.ui.interrupt_pushButton.clicked.connect(self.interrupt)
        self.settings.save_h5.connect_to_widget(self.ui.save_h5_checkBox)
        self.settings.save_movie.connect_to_widget(self.ui.save_movie_checkBox)

        # Set up pyqtgraph graph_layout in the UI
        self.graph_layout = pg.GraphicsLayoutWidget()
        self.ui.plot_groupBox.layout().addWidget(self.graph_layout)

        self.aux_graph_layout = pg.GraphicsLayoutWidget()
        self.ui.aux_plot_groupBox.layout().addWidget(self.aux_graph_layout)

        self.camera_layout = pg.GraphicsLayoutWidget()
        self.ui.camera_groupBox.layout().addWidget(self.camera_layout)

        # Create PlotItem object (a set of axes)

        self.plot1 = self.graph_layout.addPlot(row=1, col=1, title="Lick")
        self.plot2 = self.graph_layout.addPlot(row=2, col=1, title="breathing")

        # Create PlotDataItem object ( a scatter plot on the axes )
        self.breathing_plot = self.plot2.plot([0])
        self.lick_plot_0 = self.plot1.plot([0])
        self.lick_plot_1 = self.plot1.plot([1])

        self.lick_plot_0.setPen('y')
        self.lick_plot_1.setPen('g')

        self.T = np.linspace(0, 10, 10000)
        self.k = 0

        self.camera_view = pg.ViewBox()
        self.camera_layout.addItem(self.camera_view)
        self.camera_image = pg.ImageItem()
        self.camera_view.addItem(self.camera_image)

    def update_display(self):
        """
        Displays (plots) the numpy array self.buffer.
        This function runs repeatedly and automatically during the measurement run.
        its update frequency is defined by self.display_update_period
        """
        self.lick_plot_0.setData(self.k+self.T,self.buffer[:,1])
        self.lick_plot_1.setData(self.k+self.T,self.buffer[:,2])
        self.breathing_plot.setData(self.k+self.T,self.buffer[:,0])

        if self.settings.movie_on.value():
            self.camera_image.setImage(self.camera.read())
            if self.settings.save_movie.value():
                self.camera.write()


    def run(self):
        try:
            os.makedirs(self.events_path)
        except OSError:
            print("the directory %s existed" % self.events_path)
        else:
            print("created the directory %s" % self.events_path)
        logname = self.filename[0:-5] + '.log'
        logging.basicConfig(filename=logname, level=logging.DEBUG)
        logging.info(self.__dict__)
        odors_cue = OdorGen(self.list)
        odors_cue.assign_odor()
        self.reward_odor, self.non_reward_odor, = odors_cue.set_rewardodor(index=self.odor_index)
        odors_cue.initiate()
        # odors_cue.odors_DAQ[i]
        print('odor done')
        print(self.reward_odor)

        # self.waterR = DAQSimpleDOTask('Dev1/port1/line{}'.format(self.waterline))
        self.waterR = DAQSimpleDOTask('Dev3/port1/line0')
        self.waterR.low()
        # self.OdorOnCopy = DaqSimpleDOTask('dev/port/line')
        # self.OdorOnCopy.low()
        self.lickR = DAQSimpleDITask('Dev3/port1/line1')
        print('water done')

        # create excel workbook
        self.wb1 = Workbook()
        self.ws1 = self.wb1.active

        # generate trial type

        # trialtypes = np.zeros(self.numtrials)
        # for i in range(int(self.numtrials/20)):
        for i in range(self.number_of_blocks * (self.numtrials + int(round((self.numtrials) / 2)))):

            # train_nogo = np.ones(int(round(self.numtrials * 0.5)))#code 1
            # train_go = np.zeros(int(round(self.numtrials * 0.5)))#code 0

            # temp_comb1 = np.concatenate((train_go, train_nogo))
            # train_go_omission = np.ones(int(round(self.numtrials * self.p_go * (1 - self.p_reward_go)))) * 3  # code 3
            # temp_comb1 = np.concatenate((temp_comb1, train_go_omission))

            # if self.phase == 'free_water':

            # temp_comb3 = np.ones(self.numtrials) * 4

            #  train_empty = np.ones(int(self.numtrials * self.p_empty * (1 - self.p_reward_empty))) * 2
            #  temp_comb1 = np.concatenate((temp_comb1, train_empty))

            if self.phase == 'lick_train':
                temp_comb1 = np.ones(self.numtrials) * 5

            elif self.phase == 'nogo_odor_no_delay':
                # water should be available right upon licking during odor
                # should be self.duration_odor_to_action = 0
                # temp_comb1 = np.ones(int(round(1)))
                temp_comb1 = np.ones(self.numtrials)

            elif self.phase == 'go_odor_no_delay':
                # water should be available right upon licking during odor
                # should be self.duration_odor_to_action = 0
                temp_comb1 = np.zeros(self.numtrials)

            # elif self.phase == 'nogo_go_nogo_training':
            # temp_comb1 = np.concatenate(np.ones(self.numtrials/2), np.zeros(self.numtrials),np.ones(self.numtrials/2))

            elif self.phase == 'odor go_no-go_no_delay':
                # self.duration_odor_to_action = 0
                temp_comb1 = np.concatenate((train_go, train_nogo))

            elif self.phase == 'alternating blocks go_nogo':
                go = np.zeros(self.numtrials)
                no_go = np.ones(int(round(self.numtrials) / 2))
                single_block = np.concatenate((no_go, go))
                temp_comb3 = np.tile(single_block, self.number_of_blocks)






            elif self.phase == 'delayed reward':
                # they can lick as soon as odor on, but reward comes later
                # need to change go trials so water comes on later but licks counter wihout delay
                self.duration_odor_to_action = 2
                temp_comb1 = np.concatenate((train_go, train_nogo))

            elif self.phase == 'delayed licking':
                # checking period should be only after odor stopped, if lick before add delay of 5s to ITI
                # should check licks while odor on and
                # should check licks in lick period
                # get reward only if lick in right period and not wrong
                # add to ITI if lick in wrong period
                self.duration_odor_to_action = 2
                # train_empty = np.ones(self.numtrials * self.p_empty * (1-self.p_reward_empty))) * 2 #what and why
                train_go = np.ones(int(round(self.numtrials * self.p_reward_go * self.p_reward_go))) * 8
                temp_comb1 = np.concatenate((train_go, train_nogo))
                temp_comb1 = np.concatenate((temp_comb1, train_empty))

            # temp_comb2 = temp_comb1.copy()
            # random.shuffle(temp_comb2)
            # trialtypes[20*i:20*(i+1)] = temp_comb2
            trialtypes = temp_comb3
        self.trialstype = trialtypes

        print('================== Trial Types =================')
        print(trialtypes)
        t0 = time.time()
        for t in range(0, len(temp_comb3)):
            # for t in range (0, len(temp_comb1)):
            t1 = time.time()
            print('================================================')
            print('trial number: ', t)
            print('time: ', t1 - t0)

            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=101)  # 101)  # trial start
            d = self.ws1.cell(row=self.ws1.max_row, column=3, value='trial{}'.format(int(trialtypes[t])))  # trial type
            # code: 0--go w rward; 1--no go; 2--empty; 3--go w/o reward; 4 --unpred water; 5 -- c control odor； 6 -- c control odor w/o

            self.determine_reward_lickUnlimited(self.duration_rec_on_before, True)

            self.run_trial_type(int(trialtypes[t]))

            self.determine_reward_lickUnlimited(self.duration_rec_on_after, True)

            self.determine_reward_lickUnlimited(self.duration_ITI[t], True)

            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=100)
            self.wb1.save(self.filename)

        odors_cue.initiate()
        odors_cue.close()
        self.waterR.low()
        self.waterR.close()
        print('FINISHED ASSOCIATION TRAINING')

    def determine_reward_lickUnlimited(self, action_interval, check_action):
        checkperiod = 0.005
        action_timeout = time.time() + action_interval
        right_lick_last = 0
        count = 0
        reward_on = True

        while time.time() < action_timeout:
            right_lick = self.lickR.read()
            if right_lick != right_lick_last:
                if right_lick:
                    print('Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                    if check_action:
                        count += 1
                else:
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            else:
                pass
            right_lick_last = right_lick
            time.sleep(checkperiod)

        # while time.time() < action_timeout:
        # if self.punish is True and count >= self.allowedExtraLicks:
        # print('punishing ITI')
        # time.sleep(5)

        if self.operant and count < self.licknum:
            reward_on = False
            self.stat_rec.increment('miss',self.j)
        elif not self.operant:
            print('free water')
        elif self.operant and count >= self.licknum:
            self.stat_rec.increment('success',self.j)
            reward_on  = True
        reward_on

    def punishing_timeout(self, action_interval, punish):
        checkperiod = 0.005
        action_timeout = time.time() + action_interval
        right_lick_last = 0
        count = 0
        reward_on = True

        while time.time() < action_timeout:
            right_lick = self.lickR.read()
            if right_lick != right_lick_last:
                if right_lick:
                    print('Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                    if punish:
                        count += 1
                else:
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            else:
                pass
            right_lick_last = right_lick
            time.sleep(checkperiod)

        if self.punish is True and count > self.allowedExtraLicks:
            print('punishing ITI')
            time.sleep(15)

    def determine_reward_lickLimited(self, pre_action_interval, action_interval, check_action=False):
        checkperiod = 0.005

        pre_action_timeout = time.time() + pre_action_interval
        action_timeout = time.time() + action_interval

        reward_on = True
        right_lick_last = 0
        wrong_lick_last = 0

        count = 0

        while time.time() < pre_action_timeout:
            wrong_lick = self.lickR.read()
            if wrong_lick != wrong_lick_last:
                if wrong_lick:
                    print('Extraneous Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                    if check_action:
                        count += 1

                else:
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            else:
                pass
            wrong_lick_last = wrong_lick
            time.sleep(checkperiod)

        while time.time() < action_timeout:
            right_lick = self.lickR.read()
            if right_lick != right_lick_last:
                if right_lick:
                    print('Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                    if check_action:
                        count += 1

                else:
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            else:
                pass
            right_lick_last = right_lick
            time.sleep(checkperiod)

        while time.time():
            if check_action and (right_lick >= self.licknum) and (wrong_lick == 0):

                print('correct licking activate reward')
                return reward_on

            elif check_action and right_lick < self.licknum:
                print('not enough licking for reward')
                reward_on = False

            elif check_action and wrong_lick > 0:
                print('extraneous licking so no reward')
                reward_on = False
                ### need to increase ITI by 5s

    def run_trial_type(self, types):
        odor_on = False
        reward_on = False
        is_control = False

        if types == 0:
            print('go_odor_no_delay' + str(int(self.counter[types])))
            odor_on = True
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            reward_on = self.determine_reward_lickUnlimited(self.duration_odor_on + self.duration_action_window, True)
            self.run_reward_module(reward_on, w_code)

        elif types == 4:
            print('free water no odor' + str(int(self.counter[types])))
            # should not require licking
            w_code = [51, 50]
            # reward_on = True
            # self.determine_reward_lickUnlimited(self.duration_odor_on)
            # reward_
            # self.determine_reward_lickUnlimited(self.duration_odor_to_action)
            total_duration = self.duration_rec_on_before + self.duration_action_window + self.duration_water_large + self.duration_rec_on_after
            self.determine_reward_lickUnlimited(total_duration, False)
            self.run_reward_module(True, w_code)

        if types == 5:
            print('lick training no odor' + str(int(self.counter[types])))
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            duration_total = self.duration_rec_on_before + self.duration_action_window + self.duration_water_large + self.duration_rec_on_after
            reward_on = self.determine_reward_lickUnlimited(duration_total, check_action=True)
            self.run_reward_module(reward_on, w_code)

        elif types == 8:
            print('delayed lick go-odor' + str(int(self.counter[types])))
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.check_licking_1spout(self.duration_odor_on)

        elif types == 1:
            print('no go trial' + str(int(self.counter[types])))
            odor_on = True
            is_go = False
            r_code = [141, 140]
            w_code = [61, 60]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            # self.determine_reward_lickUnlimited(self.duration_odor_to_action + self.duration_action_window, True)
            reward_on = False
            self.run_reward_module(reward_on, w_code)
            self.punishing_timeout(
                self.duration_rec_on_after + self.duration_odor_on + self.duration_rec_on_before + self.duration_action_window + self.duration_water_large,
                True)
            # self.punishing_timeout((self.duration_ITI[t], True))



        elif types == 2:
            print('empty trial ' + str(int(self.counter[types])))
            # odor false: control odor comes
            w_code = [71, 70]  # not used
            # self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_on)
            self.check_licking_1spout(self.duration_odor_to_action + self.duration_action_window)
            reward_on = False
            self.run_reward_module(reward_on, w_code)

        elif types == 3:
            print('go omission trial ' + str(int(self.counter[types])))
            odor_on = True  # contingency odor comes
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action + self.duration_action_window)
            self.run_reward_module(reward_on, w_code)

    def run_odor_module(self, odor_on, is_go, is_control, r_code):
        if odor_on:
            print('opening odor port')
            if is_go and not is_control:
                self.reward_odor.high()
                time.sleep(self.duration_odor_on)
                self.reward_odor.low()
                print('closing odor port')

            elif not is_go and not is_control:
                self.non_reward_odor.high()
                time.sleep(self.duration_odor_on)
                self.non_reward_odor.low()
                print('closing odor port')

            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[1])
        else:
            self.determine_reward_lickUnlimited(self.duration_odor_on)

    def run_reward_module(self, reward_on, w_code):
        if reward_on is True:
            print('opening water valve')
            self.waterR.high()
            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[0])
            self.determine_reward_lickUnlimited(self.duration_water_large, True)
            print('closing water valve')
            self.waterR.low()
            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[1])
        else:
            self.determine_reward_lickUnlimited(self.duration_water_large, True)

    def save_training(self):
        with open(self.filename, 'wb') as output:
            pickle.dump(self, output, pickle.HIGHEST_PROTOCOL)


test = AbbyTraining()
print('start')
test.run()



