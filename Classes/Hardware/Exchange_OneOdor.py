from ScopeFoundry import Measurement
import datetime
import numpy as np
import random
import pickle
import time
import os
import logging
from Daq_Functions.DAQSimpleDOTask import DAQSimpleDOTask
from Daq_Functions.DAQSimpleDITask import DAQSimpleDITask
from Classes.Cheatproof_Define_Odors import OdorGen
from openpyxl import Workbook

class AbbyTraining(Measurement):



    def __init__(self):

        self.mouse = 'Young1'

        #self.phase='go_odor_no_delay_free20%'
        #self.phase='free_water'
        #self.phase = 'alternating blocks go_nogo'
        self.phase='go_odor_no_delay'
        #self.phase = 'nogo_odor_no_delay'
        #self.phase='cheat_go_no'
        self.condition = 'Operant'
        self.punishment = 'Punish'
        self.numtrials = 20
        self.number_of_blocks = 10

       # list of odors
        self.list = [0,3] #[nogo, nogo, nogo, go, go, go] lines
        #self.list = [6,4,2,6,4,2]  # [nogo, nogo, nogo, go, go, go] lines
        # self.list = [0, 7]
        #self.list=[0,5]
        #self.list = [0, 6]
        #self.list = [0, 4]
        #self.list = [0, 0]


        self.events_path = "C:/Abby Behavior Data/Pilot4/experiment_data_2022_3_{0}/{1}/".format(self.phase, self.mouse)
        self.events_filename = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M")+'{}.xlsx'.format(self.phase)
        # self.odor_index = [1,0]
        self.odor_index = [1,0] #odor list index POSITION 0 is reward,  1 is unrewarded, 2 is control odors

        self.allowedExtraLicks = 0

        if self.punishment == 'Punish':
           self.punish = True

        if self.condition == 'Operant':
            self.operant = True
            self.licknum = 1
            self.baseline = 0
            # make a baseline licks?
        else:
            self.operant = False
            self.licknum = 0

        # pre-training
        self.p_go = 0.4
        self.p_reward_go = 0.75
        self.p_no_go = 0.2
        self.p_empty  = 0.4
        self.p_reward_empty = 0.75
        self.cheat_odor = DAQSimpleDOTask('Dev3/port0/line7')

       # if self.phase == 'free_water':
           # self.p_reward_empty = 0
        #else:
          #  self.p_reward_empty = 0.75

        self.counter = np.zeros(9)

        self.odor_actuate = 1.2
        self.duration_rec_on_before = 2
        self.duration_odor_on = 1
        self.duration_odor_to_action = 0
        self.duration_action_window = 2
        #self.duration_action_window = 3
        self.duration_water_large = 0.03 #water
        #self.duration_water_large = 0.02 #milk
        self.duration_rec_on_after = 4
        self.duration_ITI = np.random.exponential(2, size=(self.number_of_blocks * (self.numtrials + int(round((self.numtrials) / 2)))))
        self.duration_total = self.duration_rec_on_before + self.duration_odor_on + self.duration_odor_to_action + self.duration_action_window+self.duration_water_large+self.duration_rec_on_after

        self.waterline = 3
        self.filename = self.events_path + self.events_filename

    def run(self):
        try:
            os.makedirs(self.events_path)
        except OSError:
            print("the directory %s existed" % self.events_path)
        else:
            print("created the directory %s" % self.events_path)
        logname = self.filename[0:-5] + '.log'
        logging.basicConfig(filename=logname, level=logging.DEBUG)
        logging.info(self.__dict__)
        odors_cue = OdorGen(self.list)
        odors_cue.assign_odor()
        self.reward_odor1, self.reward_odor2, self.reward_odor3, self.non_reward_odor1, self.non_reward_odor2, self.non_reward_odor3 = odors_cue.set_rewardodor(index=self.odor_index)
        self.dictionary={self.reward_odor1:str(self.list[3]), self.reward_odor2:str(self.list[4]), self.reward_odor3:str(self.list[5]),
                         self.non_reward_odor1:str(self.list[0]), self.non_reward_odor2:str(self.list[1]), self.non_reward_odor3:str(self.list[2])}
        odors_cue.initiate()
        #odors_cue.odors_DAQ[i]
        print('odor done')
        #print(self.reward_odor)

        #self.waterR = DAQSimpleDOTask('Dev1/port1/line{}'.format(self.waterline))
        self.waterR = DAQSimpleDOTask('Dev3/port1/line0') #water
        #self.waterR = DAQSimpleDOTask('Dev3/port2/line0') #milk
        self.waterR.low()
        #self.OdorOnCopy = DaqSimpleDOTask('dev/port/line')
        #self.OdorOnCopy.low()
        self.lickR = DAQSimpleDITask('Dev3/port1/line1') #water
        #self.lickR = DAQSimpleDITask('Dev3/port1/line3') #milk
        print('water done')
        self.odor_to_mouse = DAQSimpleDOTask('Dev3/port1/line2')
        self.air_to_exhaust=  DAQSimpleDOTask('Dev3/port2/line0')

        #create excel workbook
        self.wb1=Workbook()
        self.ws1=self.wb1.active

        #generate trial type

        #trialtypes = np.zeros(self.numtrials)
        #for i in range(int(self.numtrials/20)):
        for i in range (self.number_of_blocks*(self.numtrials + int(round((self.numtrials)/2)))):

            train_nogo = np.ones(int(round(self.numtrials * 0.5)))#code 1
            train_go = np.zeros(int(round(self.numtrials * 0.5)))#code 0


            #temp_comb1 = np.concatenate((train_go, train_nogo))
            #train_go_omission = np.ones(int(round(self.numtrials * self.p_go * (1 - self.p_reward_go)))) * 3  # code 3
            #temp_comb1 = np.concatenate((temp_comb1, train_go_omission))

            if self.phase == 'free_water':

               temp_comb3 = np.ones(self.numtrials) * 4

               #train_empty = np.ones(int(self.numtrials * self.p_empty * (1 - self.p_reward_empty))) * 2
               #temp_comb3 = np.concatenate((temp_comb3, train_empty))


            elif self.phase == 'nogo_odor_no_delay':
                # water should be available right upon licking during odor
                # should be self.duration_odor_to_action = 0
                #temp_comb1 = np.ones(int(round(1)))
                temp_comb3 = np.ones(self.numtrials)

            elif self.phase == 'go_odor_no_delay_free20%':
                # water should be available right upon licking during odor
                # should be self.duration_odor_to_action = 0
                go = np.zeros(int(round(((self.numtrials*0.8)))))
                free = np.ones((int(round(self.numtrials*0.2))))*4
                single_block = np.concatenate((go, free))
                np.random.shuffle(single_block)
                temp_comb3 = np.tile(single_block, self.number_of_blocks)

            elif self.phase == 'go_odor_no_delay':
                # water should be available right upon licking during odor
                # should be self.duration_odor_to_action = 0
                temp_comb3 = np.zeros(self.numtrials)

            #elif self.phase == 'nogo_go_nogo_training':
               # temp_comb1 = np.concatenate(np.ones(self.numtrials/2), np.zeros(self.numtrials),np.ones(self.numtrials/2))

            elif self.phase == 'odor go_no-go_no_delay':
                self.duration_odor_to_action = 0
                single_block = np.concatenate((train_go, train_nogo))
                temp_comb3 = np.tile(single_block, self.number_of_blocks)

            elif self.phase == 'cheat_go_no':
                self.duration_odor_to_action = 0
                number_nogo = np.ones(int(round(self.numtrials * 0.425)))  # code 1
                number_go = np.zeros(int(round(self.numtrials * 0.425)))  # code 0
                number_cheat = np.ones(int(round(self.numtrials * 0.15)))*7
                single_block =np.concatenate((number_go, number_cheat, number_nogo))
                temp_comb3  = np.tile(single_block, self.number_of_blocks)

            elif self.phase == 'alternating blocks go_nogo':
                go = np.zeros(self.numtrials)
                no_go = np.ones(int(round(self.numtrials)/2))
                single_block = np.concatenate((go, no_go))
                temp_comb3 = np.tile(single_block, self.number_of_blocks)






            elif self.phase == 'delayed reward':
                # they can lick as soon as odor on, but reward comes later
                # need to change go trials so water comes on later but licks counter wihout delay
                self.duration_odor_to_action = 2
                temp_comb1 = np.concatenate((train_go, train_nogo))

            elif self.phase == 'delayed licking':
                # checking period should be only after odor stopped, if lick before add delay of 5s to ITI
                # should check licks while odor on and
                # should check licks in lick period
                # get reward only if lick in right period and not wrong
                # add to ITI if lick in wrong period
                self.duration_odor_to_action = 2
                #train_empty = np.ones(self.numtrials * self.p_empty * (1-self.p_reward_empty))) * 2 #what and why
                train_go = np.ones(int(round(self.numtrials *self.p_reward_go*self.p_reward_go))) * 8
                temp_comb1 = np.concatenate((train_go, train_nogo))
                temp_comb1 = np.concatenate ((temp_comb1, train_empty))

            #temp_comb2 = temp_comb1.copy()
            random.shuffle(temp_comb3)
            #trialtypes[20*i:20*(i+1)] = temp_comb2
            trialtypes = temp_comb3
        self.trialstype = trialtypes

        print('================== Trial Types =================')
        print(trialtypes)
        t0 = time.time()
        for t in range (0, len(temp_comb3)):
        #for t in range (0, len(temp_comb1)):
            t1 = time.time()
            print('================================================')
            print('trial number: ', t)
            print('time: ', t1-t0)

            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value= 101) #101)  # trial start
            d = self.ws1.cell(row=self.ws1.max_row, column=3,value='trial{}'.format(int(trialtypes[t])))  # trial type
            # code: 0--go w rward; 1--no go; 2--empty; 3--go w/o reward; 4 --unpred water; 5 -- c control odor； 6 -- c control odor w/o

            self.determine_reward_lickUnlimited(self.duration_rec_on_before, True)

            #self.track_licks(self.duration_total)


            self.run_trial_type(int(trialtypes[t]))

            self.determine_reward_lickUnlimited(self.duration_rec_on_after, True)

            self.determine_reward_lickUnlimited(self.duration_ITI[t], True)
            #self.determine_reward_lickUnlimited(self.duration_ITI, True)

            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=100)
            self.wb1.save(self.filename)

        odors_cue.initiate()
        odors_cue.close()
        self.waterR.low()
        self.waterR.close()
        print ('FINISHED ASSOCIATION TRAINING')

    def track_licks(self, action_interval):
        checkperiod = 0.005
        action_timeout = time.time() + action_interval
        right_lick_last = 0


        while time.time() < action_timeout:
            right_lick = self.lickR.read()
            if right_lick != right_lick_last:
                if right_lick:
                    print('Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)
                else:
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            right_lick_last = right_lick
            time.sleep(checkperiod)


    def determine_reward_lickUnlimited(self, action_interval, check_action):
        checkperiod = 0.005
        action_timeout = time.time() + action_interval
        right_lick_last = 0
        count = 0
        reward_on = True


        while time.time() < action_timeout:
            right_lick = self.lickR.read()
            if right_lick != right_lick_last:
                if right_lick:
                    print('Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                    if check_action:
                        count += 1
                else:
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            else:
                pass
            right_lick_last = right_lick
            time.sleep(checkperiod)

       # while time.time() < action_timeout:
        #if self.punish is True and count >= self.allowedExtraLicks:
           # print('punishing ITI')
           # time.sleep(5)


        if self.operant and count < self.licknum:
            reward_on = False
        elif not self.operant:
            print('free water')
        return reward_on

    def punishing_timeout (self,action_interval, punish):
        checkperiod = 0.005
        action_timeout = time.time() + action_interval
        right_lick_last = 0
        count = 0
        reward_on = True

        while time.time() < action_timeout:
            right_lick = self.lickR.read()
            if right_lick != right_lick_last:
                if right_lick:
                    print('Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                    if punish:
                        count += 1
                else:
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            else:
                pass
            right_lick_last = right_lick
            time.sleep(checkperiod)

        if self.punish is True and count > self.allowedExtraLicks:
            print('punishing ITI')
            time.sleep(15)



    def determine_reward_lickLimited(self, pre_action_interval, action_interval, check_action=False):
        checkperiod = 0.005

        pre_action_timeout = time.time() + pre_action_interval
        action_timeout = time.time() + action_interval

        reward_on = True
        right_lick_last = 0
        wrong_lick_last = 0

        count = 0

        while time.time() < pre_action_timeout:
            wrong_lick = self.lickR.read()
            if wrong_lick != wrong_lick_last:
                if wrong_lick:
                    print('Extraneous Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                    if check_action:
                        count += 1

                else:
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            else:
                pass
            wrong_lick_last = wrong_lick
            time.sleep(checkperiod)

        while time.time() < action_timeout:
            right_lick = self.lickR.read()
            if right_lick != right_lick_last:
                if right_lick:
                    print('Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                    if check_action:
                        count += 1

                else:
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            else:
                pass
            right_lick_last = right_lick
            time.sleep(checkperiod)

        while time.time():
            if check_action and (right_lick >= self.licknum) and (wrong_lick == 0):

                print('correct licking activate reward')
                return reward_on

            elif check_action and right_lick < self.licknum:
                print('not enough licking for reward')
                reward_on = False

            elif check_action and wrong_lick > 0:
                print('extraneous licking so no reward')
                reward_on = False
                ### need to increase ITI by 5s

    def run_trial_type(self, types):
        odor_on = False
        reward_on = False
        is_control = False

        if types == 0:
            print('go_odor_no_delay'+ str(int(self.counter[types])))
            odor_on = True
            is_go = True
            is_cheat = False
            r_code=[131, 130]
            w_code = [51, 50]
            #self.track_licks(self.duration_odor_on)
            self.run_odor_module(odor_on, is_go, is_control, is_cheat, r_code)
            #reward_on = self.determine_reward_lickUnlimited(self.duration_odor_on + self.duration_action_window,  True)
            reward_on = self.determine_reward_lickUnlimited(self.duration_action_window, True)
            self.run_reward_module(reward_on, w_code)

        elif types == 7:
            print ('go_blank_cheat'+ str(int(self.counter[types])))
            odor_on = True
            is_go = False
            is_cheat = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go,  is_control, is_cheat, r_code)
            reward_on = self.determine_reward_lickUnlimited(self.duration_action_window, True)
            self.run_reward_module(reward_on, w_code)


        elif types == 4:
            print('free water no odor' + str(int(self.counter[types])))
            # should not require licking
            w_code = [51, 50]
            #reward_on = True
            #self.determine_reward_lickUnlimited(self.duration_odor_on)
            #reward_
            #self.determine_reward_lickUnlimited(self.duration_odor_to_action)
            total_duration = self.duration_rec_on_before + self.duration_action_window + self.duration_water_large + self.duration_rec_on_after
            self.determine_reward_lickUnlimited(total_duration, False)
            self.run_reward_module(True, w_code)

        if types == 5:
            print('lick training no odor' + str(int(self.counter[types])))
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            duration_total = self.duration_rec_on_before + self.duration_action_window + self.duration_water_large + self.duration_rec_on_after
            reward_on = self.determine_reward_lickUnlimited(duration_total, check_action=True)
            self.run_reward_module(reward_on, w_code)

        elif types == 8:
            print ('delayed lick go-odor' + str(int(self.counter[types])))
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.check_licking_1spout(self.duration_odor_on)

        elif types ==1:
            print ('no go trial' + str(int(self.counter[types])))
            odor_on = True
            is_go = False
            is_cheat  = False
            r_code = [141, 140]
            w_code = [61, 60]
            self.run_odor_module(odor_on, is_go, is_control, is_cheat, r_code)
            #self.determine_reward_lickUnlimited(self.duration_odor_to_action + self.duration_action_window, True)
            reward_on = False
            #self.run_reward_module(reward_on, w_code)
            #self.punishing_timeout(self.duration_rec_on_after + self.duration_odor_on +self.duration_rec_on_before + self.duration_action_window + self.duration_water_large, True)
            #self.punishing_timeout((self.duration_ITI[t], True))
            #self.punishing_timeout(self.duration_odor_on + self.duration_action_window, True)
            self.punishing_timeout(self.duration_action_window, True)




        elif types == 2:
            print('empty trial ' + str(int(self.counter[types])))
                  # odor false: control odor comes
            w_code = [71, 70]  # not used
                  # self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_on)
            self.check_licking_1spout(self.duration_odor_to_action + self.duration_action_window)
            reward_on = False
            self.run_reward_module(reward_on, w_code)

        elif types == 3:
            print('go omission trial ' + str(int(self.counter[types])))
            odor_on = True  # contingency odor comes
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action + self.duration_action_window)
            self.run_reward_module(reward_on, w_code)



    def run_odor_module(self, odor_on, is_go, is_control, is_cheat, r_code):

        if odor_on:

            print ('actuating odor')
            if is_go and not is_control:
                self.reward_odor = random.choice([self.reward_odor1, self.reward_odor2, self.reward_odor3])
                self.reward_odor.high()
                time.sleep(self.odor_actuate)
                self.odor_to_mouse.high()
                self.air_to_exhaust.high()
                print('odor to mouse')
                d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                d = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[0])

                checkperiod = 0.005
                action_timeout = time.time() + self.duration_odor_on
                right_lick_last = 0
                count = 0
                # reward_on = True
                while time.time() < action_timeout:
                    right_lick = self.lickR.read()
                    if right_lick != right_lick_last:
                        if right_lick:
                            print('Lick')
                            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                            # if check_action:
                            # count += 1
                        else:
                            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
                    else:
                        pass
                    right_lick_last = right_lick
                    time.sleep(checkperiod)

                time.sleep(self.duration_odor_on)
                self.reward_odor.low()
                self.odor_to_mouse.low()
                self.air_to_exhaust.low()
                print('closing odor port')
                print('used odor line ',self.dictionary[self.reward_odor])
                d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                d = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[1])

            elif is_cheat:
                self.reward_odor = self.cheat_odor
                self.reward_odor.high()
                time.sleep(self.odor_actuate)
                self.odor_to_mouse.high()
                self.air_to_exhaust.high()
                print('odor to mouse')
                d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                d = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[0])

                checkperiod = 0.005
                action_timeout = time.time() + self.duration_odor_on
                right_lick_last = 0
                count = 0
                # reward_on = True
                while time.time() < action_timeout:
                    right_lick = self.lickR.read()
                    if right_lick != right_lick_last:
                        if right_lick:
                            print('Lick')
                            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                            # if check_action:
                            # count += 1
                        else:
                            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
                    else:
                        pass
                    right_lick_last = right_lick
                    time.sleep(checkperiod)

                time.sleep(self.duration_odor_on)
                self.reward_odor.low()
                self.odor_to_mouse.low()
                self.air_to_exhaust.low()
                print('closing odor port')

                d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                d = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[1])


            elif not is_go and not is_control:
                self.non_reward_odor = random.choice([self.non_reward_odor1, self.non_reward_odor2, self.non_reward_odor3])
                self.non_reward_odor.high()
                time.sleep(self.odor_actuate)
                self.odor_to_mouse.high()
                print ('odor to mouse')
                d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                d = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[0])

                checkperiod = 0.005
                action_timeout = time.time() + self.duration_odor_on
                right_lick_last = 0
                count = 0
                # reward_on = True
                while time.time() < action_timeout:
                    right_lick = self.lickR.read()
                    if right_lick != right_lick_last:
                        if right_lick:
                            print('Lick')
                            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                            # if check_action:
                            # count += 1
                        else:
                            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
                    else:
                        pass
                    right_lick_last = right_lick
                    time.sleep(checkperiod)

                time.sleep(self.duration_odor_on)
                self.non_reward_odor.low()
                self.odor_to_mouse.low()
                self.air_to_exhaust.low()
                print('closing odor port')
                print('used odor line ',self.dictionary[self.non_reward_odor])
                d=self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
                d=self.ws1.cell(row=self.ws1.max_row, column = 2, value=r_code[1])
        else:
             self.determine_reward_lickUnlimited(self.duration_odor_on)

    def run_reward_module(self, reward_on,  w_code):
        if reward_on is True:
            print('opening water valve')
            self.waterR.high()
            d=self.ws1.cell(row=(self.ws1.max_row+1), column=1, value=time.time())
            d=self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[0])
            self.determine_reward_lickUnlimited(self.duration_water_large, True)
            print('closing water valve')
            self.waterR.low()
            d=self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.time())
            d=self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[1])
        else:
            self.determine_reward_lickUnlimited(self.duration_water_large, True)
    def save_training(self):
        with open(self.filename, 'wb')as output:
            pickle.dump(self, output, pickle.HIGHEST_PROTOCOL)



test = AbbyTraining()
print('start')
test.run()